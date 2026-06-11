import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


#loading dei fogli con calamine
def _load_sheet(path: str, sheet: str, engine: str = "calamine") -> pd.DataFrame:
    '''
    carica i fogli excel utilizzando calamine (provato anche con altri ma rimane il migliore in quanto tempo)

    argomenti: 
        path: path del file excel
        sheet: nome dello sheet
        engine: engine utilizzato per leggere i fogli

    return :
        un dataframe con i fogli richiesti
    raises:
        keyError se non trova il foglio indicato nel terminale

    '''
    all_sheets = pd.read_excel(path, sheet_name=None, engine=engine, dtype=str)
    if sheet in all_sheets:
        return all_sheets[sheet]
    try:
        idx = int(sheet)
        return list(all_sheets.values())[idx]
    except (ValueError, IndexError):
        available = list(all_sheets.keys())
        raise KeyError(f"Foglio non trovato, disponibili: {available}")


def write_columns(path: str, sheet_name: str, colonne_valori: dict):
    '''
    funzione per scrivere sulle colonne specificate tramite openpyxl

    args:
        path: path del foglio
        sheet_name: sheet nel quale scrivere
        colonne_valori: dizionario che deve contenere {nomeColonna : valori da inserire}

    return:
        il foglio salvato e modificato con i record inseriti correttamente

    raise:
        ValueError se non si trova lo sheet
    '''
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        raise ValueError(f"foglio '{sheet_name}'non trovato. Trovati: {wb.sheetnames}")

    headers = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[str(cell_value).strip()] = col

    for col_name, valori in colonne_valori.items():
        if col_name in headers:
            col_idx = headers[col_name]
        else:
            col_idx = max_col + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            max_col += 1
        for i, val in enumerate(valori):
            ws.cell(row=i + 2, column=col_idx, value=val if val else None)

    wb.save(path)
    print(f"{len(colonne_valori)}colonne scritte nel foglio: '{ws.title}'")


def processa_cella_societa(cella: str, societa_to_value: dict, societa_non_trovate: set, 
                           sostituzioni: dict = None) -> list:
    if not cella:
        return []
    societa_list = [s.strip() for s in cella.split("|") if s.strip()]
    values = []
    for societa in societa_list:
        if sostituzioni and societa in sostituzioni:
            societa = sostituzioni[societa]
        
        value = societa_to_value.get(societa)
        
        if not value:
            societa_lower = societa.lower()
            # a) match case‑insensitive sull'intera chiave
            for key, val in societa_to_value.items():
                if key.lower() == societa_lower:
                    value = val
                    break
            if not value and '_' not in societa:
                for key, val in societa_to_value.items():
                    if key.lower().endswith('_' + societa_lower):
                        value = val
                        break
        
        if value:
            values.append(value)
        else:
            societa_non_trovate.add(societa)
    return values
def map_societa(
    path_importtemplate: str,
    path_sip: str,
    output_sheet: str = "008",
    output_col: str = "IDS",
    output_col_conferente: str = "IDS_CONFERENTE",
    limit: int = None,
):
    
    '''
    funzione dove avviene il mapping, unisce il caricamento degli sheet di import e sip_società tramite _load_sheets(),
    la scrittura delle colonne e salvataggio dello sheet con write_columns()
    

    args:
        path_importtemplate: path del file di import
        path_sip: path del foglio dove le società sono associati ai codici alfanumerici

        output_sheet: lo sheet dove deve tornare il risultato del mapping
        output_col: la colonna dove torna il risultato
        output_col_conferente: la colonna dove torna il risultato
        
        limit: solo in fase di testing per verificare le righe limitate dal valore che si da a questo parametro

    '''
    eccezioni = {
    "0585-00_eni rete oil&nonoil S.p.A.": "0585-01_eni rete oil&nonoil",
    "0916-00":"0916-00_VERSALIS INTERNATIONAL SA",
    "Varie società":"Varie società"
    }
    print(f"\nCaricamento ImportTemplate foglio '{output_sheet}'...")
    df_output = _load_sheet(path_importtemplate, output_sheet)
    print("Caricamento file sip_societa")
    df_sip = _load_sheet(path_sip, "Sheet1")
    if limit is not None:
        df_output = df_output.head(limit)
        print(f"Limite impostato a {limit}")

    df_output.columns = df_output.columns.str.strip()
    df_sip.columns = df_sip.columns.str.strip()

    societa_to_value = {}
    for _, row in df_sip.iterrows():
        key = str(row.get("Key", "")).strip()
        value = str(row.get("Value", "")).strip()
        if key and value:
            societa_to_value[key] = value

    print(f"Trovate {len(societa_to_value)} corrispondenze di società ")

    ids_values = []
    ids_conferente_values = []
    societa_non_trovate = set()

    for idx, row in df_output.iterrows():
        societa_cell = str(row.get("societa", "")) if not pd.isna(row.get("societa")) else ""
        societa_conferente_cell = str(row.get("societa_conferente", "")) if not pd.isna(row.get("societa_conferente")) else ""

        values_societa = processa_cella_societa(societa_cell, societa_to_value, societa_non_trovate,eccezioni)
        values_conferente = processa_cella_societa(societa_conferente_cell, societa_to_value, societa_non_trovate,eccezioni)

        if values_societa:
            ids_values.append("[" + ",".join(f'"{v}"' for v in values_societa) + "]")
        else:
            ids_values.append("")
        if values_conferente:
            ids_conferente_values.append("[" + ",".join(f'"{v}"' for v in values_conferente) + "]")
        else:
            ids_conferente_values.append("")

    if societa_non_trovate:
        print(f"\nSocietà NON trovate nel file sip ({len(societa_non_trovate)}):")
        for s in sorted(societa_non_trovate)[:20]:
            print(f"  - '{s}'")
        if len(societa_non_trovate) > 20:
            print(f"e altre {len(societa_non_trovate) - 20}")

    write_columns(path_importtemplate, output_sheet, {
        output_col: ids_values,
        output_col_conferente: ids_conferente_values
    })
    print("Scrittura completata")

#Mapping MOTIVAZIONI
def processa_cella_motivazione(cella: str, motivazione_to_value: dict, 
                               motivazione_lower_to_value: dict,
                               non_trovate: set) -> list:
    if not cella:
        return []
    elementi = [e.strip() for e in cella.split("|") if e.strip()]
    valori = []
    for elemento in elementi:
        value = motivazione_to_value.get(elemento)
        if not value:
            value = motivazione_lower_to_value.get(elemento.lower())
        if value:
            valori.append(value)
        else:
            non_trovate.add(elemento)
    return valori

def map_motivazioni(
    path_importtemplate: str,
    path_sip: str,
    output_sheet: str = "025",
    output_col: str = "IDS_MOTIVAZIONE",
    output_col_conferente: str = "IDS_MOTIVAZIONI_ANNULL",
    limit: int = None,
):
    print(f"\nCaricamento ImportTemplate foglio '{output_sheet}'...")
    df_output = _load_sheet(path_importtemplate, output_sheet)
    print("Caricamento file sip_societa")
    df_sip = _load_sheet(path_sip, "Sheet1")
    if limit is not None:
        df_output = df_output.head(limit)
        print(f"Limite impostato a {limit}")

    df_output.columns = df_output.columns.str.strip()
    df_sip.columns = df_sip.columns.str.strip()

    motivazione_to_value = {}
    motivazione_lower_to_value = {}
    for _, row in df_sip.iterrows():
        key = str(row.get("Key", "")).strip()
        value = str(row.get("Value", "")).strip()
        if key and value:
            motivazione_to_value[key] = value
            if key.lower() not in motivazione_lower_to_value:
                motivazione_lower_to_value[key.lower()] = value

    print(f"Trovate {len(motivazione_to_value)} corrispondenze di motivazioni")

    def mappa_colonna(serie):
        def mappa_singola_cella(cella):
            if pd.isna(cella) or str(cella).strip() == "":
                return ""
            elementi = [e.strip() for e in str(cella).split("|") if e.strip()]
            valori = []
            for elemento in elementi:
                value = motivazione_to_value.get(elemento)
                if not value:
                    value = motivazione_lower_to_value.get(elemento.lower())
                if value:
                    valori.append(value)
            if valori:
                return "[" + ",".join(f'"{v}"' for v in valori) + "]"
            return ""
        return serie.apply(mappa_singola_cella)

    print("Mapping Motivazioni")
    ids_values = mappa_colonna(df_output.get("motivazioni", pd.Series(dtype=str)))
    ids_conferente_values = mappa_colonna(df_output.get("motivazioni_annullamento", pd.Series(dtype=str)))

    tutte_le_celle = pd.concat([
        df_output.get("motivazioni", pd.Series(dtype=str)),
        df_output.get("motivazioni_annullamento", pd.Series(dtype=str))
    ]).dropna()
    elementi_unici = set()
    for cella in tutte_le_celle:
        for e in str(cella).split("|"):
            e = e.strip()
            if e:
                elementi_unici.add(e)
    non_trovate = {e for e in elementi_unici 
                   if e not in motivazione_to_value 
                   and e.lower() not in motivazione_lower_to_value}
    if non_trovate:
        print(f"\nMotivazioni NON trovate nel file sip ({len(non_trovate)}):")
        for m in sorted(non_trovate)[:20]:
            print(f"  - '{m}'")
        if len(non_trovate) > 20:
            print(f" e altre {len(non_trovate) - 20}")

    write_columns(path_importtemplate, output_sheet, {
        output_col: ids_values.tolist(),
        output_col_conferente: ids_conferente_values.tolist()
    })
    print("Scrittura completata")


if __name__ == "__main__":
    print("Mapping tool\n")
    print("Scegliere tipo mapping: ")
    print("  1 = Società ")
    print("  2 = Motivazioni")

    scelta = input("Inserire 1 o 2: ").strip()

    importtemplate = input("Path ImportTemplate: ").strip().strip('"')
    sip = input("Path sip_societa: ").strip().strip('"')
    limit_raw = input("Limite righe per test (invio = none): ").strip()
    limit = None if not limit_raw else int(limit_raw)

    if scelta == "1":
        output_sheet = input("Foglio in ImportTemplate (invio = 008): ").strip() or "008"
        output_col = input("Colonna di output n.1 (invio = IDS): ").strip() or "IDS"
        output_col_conf = input("Colonna di output n.2 (invio = IDS_CONFERENTE): ").strip() or "IDS_CONFERENTE"
        map_societa(
            path_importtemplate=importtemplate,
            path_sip=sip,
            output_sheet=output_sheet,
            output_col=output_col,
            output_col_conferente=output_col_conf,
            limit=limit,
        )
    elif scelta == "2":
        output_sheet = input("Foglio in ImportTemplate (invio = 025): ").strip() or "025"
        output_col = input("Colonna di output n.1 (invio = IDS_MOTIVAZIONE): ").strip() or "IDS_MOTIVAZIONE"
        output_col_conf = input("Colonna di output n.2 (invio = IDS_MOTIVAZIONI_ANNULL): ").strip() or "IDS_MOTIVAZIONI_ANNULL"
        map_motivazioni(
            path_importtemplate=importtemplate,
            path_sip=sip,
            output_sheet=output_sheet,
            output_col=output_col,
            output_col_conferente=output_col_conf,
            limit=limit,
        )
    else:
        print("Scelta non valida")

    print("\nCompletato")